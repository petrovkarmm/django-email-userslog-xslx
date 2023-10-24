def run():
    import os
    from .models import Profile # inventic profile model
    from datetime import date, datetime
    from django.contrib.auth.models import User
    from openpyxl import Workbook
    from django.core.mail import EmailMessage
    from django.conf import settings

    wb = Workbook()
    ws = wb.active  # Activity list

    # EmailMessage(to=)
    admin_mail = ['adminmail@email.com']

    # server time
    current_day = date.today()
    current_date_with_time = datetime.now()
    ws.title = f"Отчёт активности{current_day}"
    columns_titles = [
        '№',
        'Email учителя',
        'Фамилия',
        'Имя',
        'Телефон',
        'Последняя активность',
        'Дата регистрации',
    ]

    email_data = '| Email не указан |'
    last_name_data = '| Фамилия не указана |'
    first_name_data = '| Имя не указано |'
    phone_data = '| Номер не указан |'
    register_data = '| Официально не зарегистрирован (?) Желательно проверить юзера |'

    # В зависимости от количества таблиц сверху, нужно будет продолжить алфавит
    # согласно excel (A1:H1 и т.п) иначе просто будет пусто
    head_cell_range = ws['A1':'G1'][0]
    head_tuples_list = list(zip(head_cell_range, columns_titles))
    for head_cell_tuple in head_tuples_list:
        cell = head_cell_tuple[0]
        cell.value = head_cell_tuple[1]

    users = User.objects.filter(profile__role=Profile.UserRoles.TEACHER).order_by('-last_login')
    for i, user in enumerate(users):
        line_num = i + 1

        user_activity_date = user.last_login.strftime("%d-%m-%Y  %H:%M")
        date_joined = user.date_joined.strftime("%d-%m-%Y  %H:%M")

        values = [
            line_num,
            user.email or email_data,
            f'{user.last_name or last_name_data}',
            f'{user.first_name or first_name_data}',
            f'{user.profile.phone or phone_data}',
            user_activity_date,
            f'{date_joined or register_data}',
        ]
        row_number = line_num + 1
        # Аналогичено head_cell_range обновить G --> H и т.п
        cell_range = ws[f'A{row_number}':f'G{row_number}'][0]
        range_with_values = list(zip(cell_range, values))
        for cell_tuple in range_with_values:
            cell = cell_tuple[0]
            cell.value = cell_tuple[1]

    wb.template = False
    wb.save(f'users_data_{current_day}.xlsx')
    # save excel

    file_name = f'users_data_{current_day}.xlsx'
    path_to_file = os.path.abspath(file_name)

    # django email sending operation
    with open(path_to_file, 'rb') as file:
        email = EmailMessage(
            subject=f'Inventic {current_day} Users Log',
            body=f'Time at which the script was launched:\n'
                 f'{current_date_with_time}\n'
                 f'(CURRENT DATE ON SERVER)',
            from_email=settings.EMAIL_HOST_USER,
            to=admin_mail,
        )

        email.attach('file.xlsx',
                     file.read(),
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        email.send()


if __name__ == '__main__' or __name__ == 'django.core.management.commands.shell':
    run()